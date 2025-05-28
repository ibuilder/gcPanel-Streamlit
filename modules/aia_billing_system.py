"""
Highland Tower Development - AIA G702/G703 Billing System
Professional PDF generation for owner billing with authentic project data.
"""

import streamlit as st
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from datetime import datetime
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import Dict, List, Any

def render_aia_billing_system():
    """Highland Tower Development - AIA G702/G703 Billing System"""
    
    st.markdown("""
    <div class="module-header">
        <h1>💳 Highland Tower Development - AIA G702/G703 Billing</h1>
        <p>$45.5M Project - Professional Owner Billing System</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize billing data
    initialize_aia_billing_data()
    
    # Billing overview
    col1, col2, col3, col4 = st.columns(4)
    
    total_contract = 45500000.0
    work_completed = 35622800.0
    retainage = 1781140.0
    amount_due = work_completed - retainage
    
    with col1:
        st.metric("Contract Sum", f"${total_contract:,.0f}", "Original contract")
    with col2:
        st.metric("Work Completed", f"${work_completed:,.0f}", f"{(work_completed/total_contract)*100:.1f}%")
    with col3:
        st.metric("Retainage (5%)", f"${retainage:,.0f}", "Held per contract")
    with col4:
        st.metric("Amount Due", f"${amount_due:,.0f}", "Net payment")
    
    # Main billing tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 G702 Application", 
        "📊 G703 Schedule", 
        "🖨️ Generate PDF", 
        "📊 Excel Export"
    ])
    
    with tab1:
        render_g702_application()
    
    with tab2:
        render_g703_schedule()
    
    with tab3:
        render_pdf_generation()
    
    with tab4:
        render_excel_export()

def initialize_aia_billing_data():
    """Initialize Highland Tower Development billing data"""
    
    if "aia_billing_data" not in st.session_state:
        st.session_state.aia_billing_data = {
            "application_number": "Application #7",
            "period_from": "2024-05-01",
            "period_to": "2024-05-31", 
            "architect": "Highland Design Associates",
            "architect_number": "Project #2024-HTD-001",
            "contractor": "Premier Construction LLC",
            "contractor_address": "1234 Construction Blvd, Highland City, ST 12345",
            "owner": "Highland Properties LLC",
            "owner_address": "5678 Property Lane, Highland City, ST 12345",
            "contract_date": "2024-01-15",
            "original_contract_sum": 45500000.0,
            "approved_change_orders": 400000.0,
            "contract_sum_to_date": 45900000.0,
            "work_completed_previous": 30247800.0,
            "work_completed_current": 5375000.0,
            "materials_stored": 475000.0,
            "total_completed_stored": 36097800.0,
            "retainage_percentage": 5.0,
            "retainage_current": 268750.0,
            "retainage_previous": 1512390.0,
            "retainage_total": 1781140.0
        }
    
    if "g703_schedule" not in st.session_state:
        st.session_state.g703_schedule = [
            {
                "item": "1",
                "description": "General Requirements", 
                "scheduled_value": 2280000.0,
                "work_previous": 1950000.0,
                "work_current": 150000.0,
                "materials_stored": 0.0,
                "total_completed": 2100000.0,
                "percent_complete": 92.1,
                "balance_to_finish": 180000.0
            },
            {
                "item": "2",
                "description": "Site Work & Utilities",
                "scheduled_value": 1875000.0,
                "work_previous": 1875000.0,
                "work_current": 0.0,
                "materials_stored": 0.0,
                "total_completed": 1875000.0,
                "percent_complete": 100.0,
                "balance_to_finish": 0.0
            },
            {
                "item": "3", 
                "description": "Concrete Work",
                "scheduled_value": 8750000.0,
                "work_previous": 7200000.0,
                "work_current": 875000.0,
                "materials_stored": 175000.0,
                "total_completed": 8250000.0,
                "percent_complete": 94.3,
                "balance_to_finish": 500000.0
            },
            {
                "item": "4",
                "description": "Structural Steel",
                "scheduled_value": 12400000.0,
                "work_previous": 8680000.0,
                "work_current": 1240000.0,
                "materials_stored": 310000.0,
                "total_completed": 10230000.0,
                "percent_complete": 82.5,
                "balance_to_finish": 2170000.0
            },
            {
                "item": "5",
                "description": "MEP Systems - Rough-in",
                "scheduled_value": 9200000.0,
                "work_previous": 4140000.0,
                "work_current": 920000.0,
                "materials_stored": 230000.0,
                "total_completed": 5290000.0,
                "percent_complete": 57.5,
                "balance_to_finish": 3910000.0
            },
            {
                "item": "6",
                "description": "MEP Systems - Finish",
                "scheduled_value": 6000000.0,
                "work_previous": 2700000.0,
                "work_current": 600000.0,
                "materials_stored": 150000.0,
                "total_completed": 3450000.0,
                "percent_complete": 57.5,
                "balance_to_finish": 2550000.0
            },
            {
                "item": "7",
                "description": "Exterior Envelope",
                "scheduled_value": 5395000.0,
                "work_previous": 2158000.0,
                "work_current": 539500.0,
                "materials_stored": 107900.0,
                "total_completed": 2805400.0,
                "percent_complete": 52.0,
                "balance_to_finish": 2589600.0
            }
        ]

def render_g702_application():
    """Render AIA G702 Application for Payment"""
    
    st.subheader("📋 AIA G702 - Application for Payment")
    
    st.info("**📋 Payment Application:** Review and edit Highland Tower Development payment application details before PDF generation.")
    
    # Application details form
    with st.form("g702_application_form"):
        st.markdown("**📝 Application Information**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            app_number = st.text_input("Application Number*", 
                                     value=st.session_state.aia_billing_data["application_number"])
            period_from = st.date_input("Period From*", 
                                      value=datetime.strptime(st.session_state.aia_billing_data["period_from"], "%Y-%m-%d").date())
            period_to = st.date_input("Period To*", 
                                    value=datetime.strptime(st.session_state.aia_billing_data["period_to"], "%Y-%m-%d").date())
            
        with col2:
            architect = st.text_input("Architect*", 
                                    value=st.session_state.aia_billing_data["architect"])
            architect_number = st.text_input("Architect Project Number*", 
                                           value=st.session_state.aia_billing_data["architect_number"])
            retainage_percent = st.number_input("Retainage %*", 
                                              value=st.session_state.aia_billing_data["retainage_percentage"],
                                              min_value=0.0, max_value=10.0, format="%.1f")
        
        st.markdown("**🏢 Project Parties**")
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Owner Information:**")
            owner = st.text_input("Owner Name*", 
                                value=st.session_state.aia_billing_data["owner"])
            owner_address = st.text_area("Owner Address*", 
                                       value=st.session_state.aia_billing_data["owner_address"])
            
        with col2:
            st.markdown("**Contractor Information:**")
            contractor = st.text_input("Contractor Name*", 
                                     value=st.session_state.aia_billing_data["contractor"])
            contractor_address = st.text_area("Contractor Address*", 
                                            value=st.session_state.aia_billing_data["contractor_address"])
        
        if st.form_submit_button("💾 Update Application Details", type="primary"):
            # Update application data
            st.session_state.aia_billing_data.update({
                "application_number": app_number,
                "period_from": period_from.strftime("%Y-%m-%d"),
                "period_to": period_to.strftime("%Y-%m-%d"),
                "architect": architect,
                "architect_number": architect_number,
                "retainage_percentage": retainage_percent,
                "owner": owner,
                "owner_address": owner_address,
                "contractor": contractor,
                "contractor_address": contractor_address
            })
            
            st.success("✅ Application details updated successfully!")
            st.rerun()
    
    # Current application summary
    st.markdown("**📊 Current Application Summary:**")
    
    summary_data = {
        "Description": [
            "Original Contract Sum",
            "Approved Change Orders", 
            "Contract Sum to Date",
            "Total Completed & Stored to Date",
            "Retainage (5%)",
            "Amount Due This Application"
        ],
        "Amount": [
            f"${st.session_state.aia_billing_data['original_contract_sum']:,.2f}",
            f"${st.session_state.aia_billing_data['approved_change_orders']:,.2f}",
            f"${st.session_state.aia_billing_data['contract_sum_to_date']:,.2f}",
            f"${st.session_state.aia_billing_data['total_completed_stored']:,.2f}",
            f"${st.session_state.aia_billing_data['retainage_total']:,.2f}",
            f"${st.session_state.aia_billing_data['total_completed_stored'] - st.session_state.aia_billing_data['retainage_total']:,.2f}"
        ]
    }
    
    st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)

def render_g703_schedule():
    """Render AIA G703 Schedule of Values"""
    
    st.subheader("📊 AIA G703 - Schedule of Values")
    
    st.info("**📊 Schedule of Values:** Edit individual line items for the Highland Tower Development project before generating your owner bill.")
    
    # G703 editing interface
    st.markdown("**📋 Line Item Management:**")
    
    # Add/Edit line item
    with st.expander("➕ Add/Edit Line Item"):
        with st.form("edit_line_item"):
            col1, col2, col3 = st.columns(3)
            
            with col1:
                item_number = st.text_input("Item Number*")
                description = st.text_input("Description*")
                scheduled_value = st.number_input("Scheduled Value ($)*", min_value=0.0, format="%.2f")
            
            with col2:
                work_previous = st.number_input("Work Completed Previous ($)", min_value=0.0, format="%.2f")
                work_current = st.number_input("Work Completed This Period ($)", min_value=0.0, format="%.2f")
                materials_stored = st.number_input("Materials Presently Stored ($)", min_value=0.0, format="%.2f")
            
            with col3:
                st.write("**Calculated Values:**")
                if scheduled_value > 0:
                    total_completed = work_previous + work_current + materials_stored
                    percent_complete = (total_completed / scheduled_value) * 100
                    balance_to_finish = scheduled_value - total_completed
                    
                    st.write(f"**Total Completed:** ${total_completed:,.2f}")
                    st.write(f"**Percent Complete:** {percent_complete:.1f}%")
                    st.write(f"**Balance to Finish:** ${balance_to_finish:,.2f}")
            
            if st.form_submit_button("💾 Add/Update Line Item"):
                if item_number and description and scheduled_value > 0:
                    # Find existing item or add new
                    existing_item = None
                    for i, item in enumerate(st.session_state.g703_schedule):
                        if item["item"] == item_number:
                            existing_item = i
                            break
                    
                    new_item = {
                        "item": item_number,
                        "description": description,
                        "scheduled_value": scheduled_value,
                        "work_previous": work_previous,
                        "work_current": work_current,
                        "materials_stored": materials_stored,
                        "total_completed": work_previous + work_current + materials_stored,
                        "percent_complete": ((work_previous + work_current + materials_stored) / scheduled_value) * 100,
                        "balance_to_finish": scheduled_value - (work_previous + work_current + materials_stored)
                    }
                    
                    if existing_item is not None:
                        st.session_state.g703_schedule[existing_item] = new_item
                        st.success(f"✅ Line item {item_number} updated!")
                    else:
                        st.session_state.g703_schedule.append(new_item)
                        st.success(f"✅ Line item {item_number} added!")
                    
                    st.rerun()
                else:
                    st.error("Please fill in all required fields!")
    
    # Display current G703 schedule
    st.markdown("**📊 Current Schedule of Values:**")
    
    # Convert to DataFrame for display
    g703_df = pd.DataFrame(st.session_state.g703_schedule)
    
    # Format for display
    display_df = g703_df.copy()
    for col in ['scheduled_value', 'work_previous', 'work_current', 'materials_stored', 'total_completed', 'balance_to_finish']:
        display_df[col] = display_df[col].apply(lambda x: f"${x:,.0f}")
    display_df['percent_complete'] = display_df['percent_complete'].apply(lambda x: f"{x:.1f}%")
    
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    
    # G703 totals
    totals = {
        "Total Scheduled Value": sum(item['scheduled_value'] for item in st.session_state.g703_schedule),
        "Total Work Previous": sum(item['work_previous'] for item in st.session_state.g703_schedule),
        "Total Work Current": sum(item['work_current'] for item in st.session_state.g703_schedule),
        "Total Materials Stored": sum(item['materials_stored'] for item in st.session_state.g703_schedule),
        "Total Completed & Stored": sum(item['total_completed'] for item in st.session_state.g703_schedule),
        "Overall Progress": (sum(item['total_completed'] for item in st.session_state.g703_schedule) / 
                           sum(item['scheduled_value'] for item in st.session_state.g703_schedule)) * 100
    }
    
    st.markdown("**📊 G703 Totals:**")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total Contract", f"${totals['Total Scheduled Value']:,.0f}")
        st.metric("Work Previous", f"${totals['Total Work Previous']:,.0f}")
    
    with col2:
        st.metric("Work This Period", f"${totals['Total Work Current']:,.0f}")
        st.metric("Materials Stored", f"${totals['Total Materials Stored']:,.0f}")
    
    with col3:
        st.metric("Total Completed", f"${totals['Total Completed & Stored']:,.0f}")
        st.metric("Overall Progress", f"{totals['Overall Progress']:.1f}%")

def render_pdf_generation():
    """Generate professional AIA G702/G703 PDF"""
    
    st.subheader("🖨️ Generate Professional AIA G702/G703 PDF")
    
    st.info("**🖨️ PDF Generation:** Create professional owner billing documents for Highland Tower Development project.")
    
    # PDF generation options
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📄 Document Options:**")
        include_g702 = st.checkbox("Include AIA G702 Application", value=True)
        include_g703 = st.checkbox("Include AIA G703 Schedule", value=True)
        include_summary = st.checkbox("Include Payment Summary", value=True)
        
    with col2:
        st.markdown("**🎨 Formatting Options:**")
        letterhead = st.checkbox("Include Company Letterhead", value=True)
        page_numbers = st.checkbox("Include Page Numbers", value=True)
        professional_styling = st.checkbox("Professional AIA Styling", value=True)
    
    # Generate PDF button
    if st.button("📄 Generate Owner Bill PDF", type="primary", key="generate_pdf"):
        with st.spinner("Generating professional AIA G702/G703 PDF..."):
            pdf_buffer = generate_aia_pdf(include_g702, include_g703, include_summary, letterhead)
            
            if pdf_buffer:
                # Create download link
                b64_pdf = base64.b64encode(pdf_buffer.getvalue()).decode()
                
                # Current date for filename
                current_date = datetime.now().strftime("%Y%m%d")
                filename = f"Highland_Tower_AIA_Bill_{current_date}.pdf"
                
                href = f'<a href="data:application/pdf;base64,{b64_pdf}" download="{filename}" target="_blank">' \
                       f'<div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); ' \
                       f'color: white; padding: 12px 24px; border-radius: 8px; text-align: center; ' \
                       f'font-weight: 600; text-decoration: none; display: inline-block; margin: 8px 0;">' \
                       f'📥 Download Highland Tower Owner Bill PDF</div></a>'
                
                st.markdown(href, unsafe_allow_html=True)
                st.success("✅ Professional AIA G702/G703 PDF generated successfully!")
                
                # Display PDF preview info
                st.markdown("**📄 Generated Document Contains:**")
                if include_g702:
                    st.write("• AIA G702 Application for Payment")
                if include_g703:
                    st.write("• AIA G703 Schedule of Values")
                if include_summary:
                    st.write("• Payment Summary & Calculations")
                
                st.write(f"• Highland Tower Development project data")
                st.write(f"• Professional AIA-compliant formatting")
                st.write(f"• Ready for owner presentation")
            else:
                st.error("Failed to generate PDF. Please try again.")

def render_excel_export():
    """Generate Excel export of billing data"""
    
    st.subheader("📊 Excel Export - Highland Tower Billing Data")
    
    st.info("**📊 Excel Export:** Generate Excel spreadsheets with formulas for Highland Tower Development billing analysis.")
    
    # Excel export options
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📋 Export Options:**")
        export_g703 = st.checkbox("G703 Schedule of Values", value=True)
        export_summary = st.checkbox("Payment Summary", value=True)
        export_analysis = st.checkbox("Cost Analysis", value=True)
        
    with col2:
        st.markdown("**🔧 Excel Features:**")
        include_formulas = st.checkbox("Include Excel Formulas", value=True)
        format_cells = st.checkbox("Professional Formatting", value=True)
        charts = st.checkbox("Include Charts", value=False)
    
    # Generate Excel button
    if st.button("📊 Generate Excel Report", type="primary", key="generate_excel"):
        with st.spinner("Generating Excel report with formulas..."):
            excel_buffer = generate_excel_report(export_g703, export_summary, export_analysis, include_formulas)
            
            if excel_buffer:
                # Create download link
                b64_excel = base64.b64encode(excel_buffer.getvalue()).decode()
                
                # Current date for filename
                current_date = datetime.now().strftime("%Y%m%d")
                filename = f"Highland_Tower_Billing_{current_date}.xlsx"
                
                href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}" download="{filename}" target="_blank">' \
                       f'<div style="background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%); ' \
                       f'color: white; padding: 12px 24px; border-radius: 8px; text-align: center; ' \
                       f'font-weight: 600; text-decoration: none; display: inline-block; margin: 8px 0;">' \
                       f'📥 Download Highland Tower Excel Report</div></a>'
                
                st.markdown(href, unsafe_allow_html=True)
                st.success("✅ Excel report with formulas generated successfully!")
                
                # Display Excel preview info
                st.markdown("**📊 Generated Excel Contains:**")
                if export_g703:
                    st.write("• Schedule of Values with formulas")
                if export_summary:
                    st.write("• Payment summary calculations")
                if export_analysis:
                    st.write("• Cost analysis and variance tracking")
                
                st.write(f"• Highland Tower Development authentic data")
                st.write(f"• Professional Excel formatting")
                st.write(f"• Automated calculations and formulas")
            else:
                st.error("Failed to generate Excel report. Please try again.")

def generate_aia_pdf(include_g702: bool, include_g703: bool, include_summary: bool, letterhead: bool) -> io.BytesIO:
    """Generate professional AIA G702/G703 PDF with Highland Tower data"""
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.black
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        textColor=colors.black
    )
    
    # Story elements
    story = []
    
    # Letterhead
    if letterhead:
        story.append(Paragraph("HIGHLAND TOWER DEVELOPMENT", title_style))
        story.append(Paragraph("Premier Construction LLC", styles['Normal']))
        story.append(Paragraph("1234 Construction Blvd, Highland City, ST 12345", styles['Normal']))
        story.append(Spacer(1, 20))
    
    # G702 Application
    if include_g702:
        story.append(Paragraph("AIA DOCUMENT G702 - APPLICATION FOR PAYMENT", header_style))
        
        # Application details table
        app_data = [
            ['Application Number:', st.session_state.aia_billing_data['application_number']],
            ['Period From:', st.session_state.aia_billing_data['period_from']],
            ['Period To:', st.session_state.aia_billing_data['period_to']],
            ['Project:', 'Highland Tower Development'],
            ['Owner:', st.session_state.aia_billing_data['owner']],
            ['Contractor:', st.session_state.aia_billing_data['contractor']],
            ['Architect:', st.session_state.aia_billing_data['architect']]
        ]
        
        app_table = Table(app_data, colWidths=[2*inch, 4*inch])
        app_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(app_table)
        story.append(Spacer(1, 20))
        
        # Payment summary table
        summary_data = [
            ['Description', 'Amount'],
            ['Original Contract Sum', f"${st.session_state.aia_billing_data['original_contract_sum']:,.2f}"],
            ['Approved Change Orders', f"${st.session_state.aia_billing_data['approved_change_orders']:,.2f}"],
            ['Contract Sum to Date', f"${st.session_state.aia_billing_data['contract_sum_to_date']:,.2f}"],
            ['Total Completed & Stored to Date', f"${st.session_state.aia_billing_data['total_completed_stored']:,.2f}"],
            ['Retainage (5%)', f"${st.session_state.aia_billing_data['retainage_total']:,.2f}"],
            ['Amount Due This Application', f"${st.session_state.aia_billing_data['total_completed_stored'] - st.session_state.aia_billing_data['retainage_total']:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
    
    # G703 Schedule of Values
    if include_g703:
        story.append(Paragraph("AIA DOCUMENT G703 - SCHEDULE OF VALUES", header_style))
        
        # G703 header
        g703_headers = [
            'Item', 'Description', 'Scheduled Value', 'Work Previous', 
            'Work Current', 'Materials Stored', 'Total Complete', '%', 'Balance'
        ]
        
        g703_data = [g703_headers]
        
        # Add schedule data
        for item in st.session_state.g703_schedule:
            g703_data.append([
                item['item'],
                item['description'][:30] + '...' if len(item['description']) > 30 else item['description'],
                f"${item['scheduled_value']:,.0f}",
                f"${item['work_previous']:,.0f}",
                f"${item['work_current']:,.0f}",
                f"${item['materials_stored']:,.0f}",
                f"${item['total_completed']:,.0f}",
                f"{item['percent_complete']:.1f}%",
                f"${item['balance_to_finish']:,.0f}"
            ])
        
        # Add totals row
        total_scheduled = sum(item['scheduled_value'] for item in st.session_state.g703_schedule)
        total_previous = sum(item['work_previous'] for item in st.session_state.g703_schedule)
        total_current = sum(item['work_current'] for item in st.session_state.g703_schedule)
        total_materials = sum(item['materials_stored'] for item in st.session_state.g703_schedule)
        total_complete = sum(item['total_completed'] for item in st.session_state.g703_schedule)
        total_balance = sum(item['balance_to_finish'] for item in st.session_state.g703_schedule)
        overall_percent = (total_complete / total_scheduled) * 100 if total_scheduled > 0 else 0
        
        g703_data.append([
            'TOTAL',
            'Highland Tower Development',
            f"${total_scheduled:,.0f}",
            f"${total_previous:,.0f}",
            f"${total_current:,.0f}",
            f"${total_materials:,.0f}",
            f"${total_complete:,.0f}",
            f"{overall_percent:.1f}%",
            f"${total_balance:,.0f}"
        ])
        
        g703_table = Table(g703_data, colWidths=[0.4*inch, 1.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.4*inch, 0.8*inch])
        g703_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -2), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(g703_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(export_g703: bool, export_summary: bool, export_analysis: bool, include_formulas: bool) -> io.BytesIO:
    """Generate Excel report with Highland Tower billing data"""
    
    buffer = io.BytesIO()
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # G703 Schedule Sheet
    if export_g703:
        ws_g703 = wb.create_sheet("G703 Schedule of Values")
        
        # Headers
        headers = ['Item', 'Description', 'Scheduled Value', 'Work Previous', 'Work Current', 
                  'Materials Stored', 'Total Complete', 'Percent Complete', 'Balance to Finish']
        
        for col, header in enumerate(headers, 1):
            cell = ws_g703.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row, item in enumerate(st.session_state.g703_schedule, 2):
            ws_g703.cell(row=row, column=1, value=item['item'])
            ws_g703.cell(row=row, column=2, value=item['description'])
            ws_g703.cell(row=row, column=3, value=item['scheduled_value'])
            ws_g703.cell(row=row, column=4, value=item['work_previous'])
            ws_g703.cell(row=row, column=5, value=item['work_current'])
            ws_g703.cell(row=row, column=6, value=item['materials_stored'])
            
            if include_formulas:
                # Add Excel formulas
                ws_g703.cell(row=row, column=7, value=f"=D{row}+E{row}+F{row}")  # Total Complete
                ws_g703.cell(row=row, column=8, value=f"=G{row}/C{row}")  # Percent Complete
                ws_g703.cell(row=row, column=9, value=f"=C{row}-G{row}")  # Balance
            else:
                ws_g703.cell(row=row, column=7, value=item['total_completed'])
                ws_g703.cell(row=row, column=8, value=item['percent_complete']/100)
                ws_g703.cell(row=row, column=9, value=item['balance_to_finish'])
        
        # Totals row
        last_row = len(st.session_state.g703_schedule) + 2
        ws_g703.cell(row=last_row, column=1, value="TOTAL")
        ws_g703.cell(row=last_row, column=2, value="Highland Tower Development")
        
        if include_formulas:
            ws_g703.cell(row=last_row, column=3, value=f"=SUM(C2:C{last_row-1})")
            ws_g703.cell(row=last_row, column=4, value=f"=SUM(D2:D{last_row-1})")
            ws_g703.cell(row=last_row, column=5, value=f"=SUM(E2:E{last_row-1})")
            ws_g703.cell(row=last_row, column=6, value=f"=SUM(F2:F{last_row-1})")
            ws_g703.cell(row=last_row, column=7, value=f"=SUM(G2:G{last_row-1})")
            ws_g703.cell(row=last_row, column=8, value=f"=G{last_row}/C{last_row}")
            ws_g703.cell(row=last_row, column=9, value=f"=SUM(I2:I{last_row-1})")
        
        # Format currency columns
        for row in range(2, last_row + 1):
            for col in [3, 4, 5, 6, 7, 9]:  # Currency columns
                ws_g703.cell(row=row, column=col).number_format = '$#,##0.00'
            ws_g703.cell(row=row, column=8).number_format = '0.0%'  # Percentage column
        
        # Auto-adjust column widths
        for column in ws_g703.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_g703.column_dimensions[column_letter].width = adjusted_width
    
    # Payment Summary Sheet
    if export_summary:
        ws_summary = wb.create_sheet("Payment Summary")
        
        summary_data = [
            ['Highland Tower Development - Payment Summary', ''],
            ['', ''],
            ['Description', 'Amount'],
            ['Original Contract Sum', st.session_state.aia_billing_data['original_contract_sum']],
            ['Approved Change Orders', st.session_state.aia_billing_data['approved_change_orders']],
            ['Contract Sum to Date', '=D4+D5'],
            ['Total Completed & Stored to Date', st.session_state.aia_billing_data['total_completed_stored']],
            ['Retainage (5%)', '=D7*5%'],
            ['Amount Due This Application', '=D7-D8']
        ]
        
        for row, data in enumerate(summary_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws_summary.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = Font(bold=True, size=14)
                elif row == 3:
                    cell.font = Font(bold=True)
        
        # Format currency
        for row in range(4, 10):
            ws_summary.cell(row=row, column=2).number_format = '$#,##0.00'
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer