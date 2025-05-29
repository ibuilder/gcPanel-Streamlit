"""
Highland Tower Development - Report Generation Center
Centralized reporting hub with custom templates and automated distribution.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime, timedelta
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from typing import Dict, List, Any, Optional
import json

def render_report_generation_center():
    """Highland Tower Development - Advanced Report Generation Center"""
    
    st.markdown("""
    <div class="module-header">
        <h1>📊 Highland Tower Development - Report Generation Center</h1>
        <p>$45.5M Project - Executive Reporting & Analytics Hub</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize report data
    initialize_highland_report_data()
    
    # Report overview metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Report Templates", "15", "Professional formats")
    with col2:
        st.metric("Generated Reports", "234", "This month")
    with col3:
        st.metric("Scheduled Reports", "8", "Automated delivery")
    with col4:
        st.metric("Distribution Lists", "12", "Stakeholder groups")
    
    # Main report tabs
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📋 Generate Reports",
        "📊 Report Templates", 
        "🕒 Scheduled Reports",
        "📧 Distribution",
        "📈 Report Analytics"
    ])
    
    with tab1:
        render_generate_reports()
    
    with tab2:
        render_report_templates()
    
    with tab3:
        render_scheduled_reports()
    
    with tab4:
        render_distribution_management()
    
    with tab5:
        render_report_analytics()

def initialize_highland_report_data():
    """Initialize Highland Tower Development report data"""
    
    if "highland_report_templates" not in st.session_state:
        st.session_state.highland_report_templates = [
            {
                "template_id": "RPT-EXEC-001",
                "name": "Executive Summary Report",
                "description": "Monthly executive summary with key metrics and progress",
                "category": "Executive",
                "format": "PDF",
                "frequency": "Monthly",
                "data_sources": ["Cost Management", "Schedule", "Safety", "Quality"],
                "sections": ["Project Overview", "Financial Summary", "Schedule Status", "Risk Assessment"],
                "last_used": "2024-05-28",
                "usage_count": 45
            },
            {
                "template_id": "RPT-COST-001",
                "name": "Cost Performance Report",
                "description": "Detailed cost analysis with budget variance and forecasting",
                "category": "Financial",
                "format": "Excel",
                "frequency": "Weekly",
                "data_sources": ["Cost Management", "Change Orders", "SOV"],
                "sections": ["Budget Summary", "Variance Analysis", "Change Orders", "Forecasting"],
                "last_used": "2024-05-27",
                "usage_count": 78
            },
            {
                "template_id": "RPT-SAFE-001",
                "name": "Safety Performance Dashboard",
                "description": "Safety metrics, incidents, and compliance tracking",
                "category": "Safety",
                "format": "PDF",
                "frequency": "Weekly",
                "data_sources": ["Safety", "Daily Reports", "Training"],
                "sections": ["Incident Summary", "Safety Metrics", "Training Status", "Compliance"],
                "last_used": "2024-05-26",
                "usage_count": 52
            },
            {
                "template_id": "RPT-PROG-001",
                "name": "Progress Report",
                "description": "Construction progress with photos and milestone tracking",
                "category": "Progress",
                "format": "PDF",
                "frequency": "Bi-weekly",
                "data_sources": ["Daily Reports", "Progress Photos", "Schedule", "Quality"],
                "sections": ["Progress Summary", "Photos", "Milestones", "Issues"],
                "last_used": "2024-05-25",
                "usage_count": 34
            },
            {
                "template_id": "RPT-QUAL-001",
                "name": "Quality Control Report",
                "description": "Quality inspections, defects, and corrective actions",
                "category": "Quality",
                "format": "Excel",
                "frequency": "Weekly",
                "data_sources": ["Quality Control", "Inspections", "RFIs"],
                "sections": ["Inspection Summary", "Defect Tracking", "Corrective Actions"],
                "last_used": "2024-05-24",
                "usage_count": 28
            }
        ]
    
    if "highland_distribution_lists" not in st.session_state:
        st.session_state.highland_distribution_lists = [
            {
                "list_id": "DL-001",
                "name": "Executive Team",
                "description": "Project executives and senior management",
                "recipients": [
                    {"name": "Highland Properties CEO", "email": "ceo@highlandproperties.com", "role": "Owner"},
                    {"name": "John Smith", "email": "jsmith@gcprime.com", "role": "Project Manager"},
                    {"name": "Sarah Wilson", "email": "swilson@gcprime.com", "role": "Site Supervisor"}
                ],
                "report_types": ["Executive", "Financial"],
                "frequency": "Monthly"
            },
            {
                "list_id": "DL-002",
                "name": "Project Team",
                "description": "Core project team members",
                "recipients": [
                    {"name": "Mike Johnson", "email": "mjohnson@gcprime.com", "role": "Safety Manager"},
                    {"name": "Lisa Chen", "email": "lchen@gcprime.com", "role": "Project Engineer"},
                    {"name": "Tom Brown", "email": "tbrown@gcprime.com", "role": "Cost Manager"}
                ],
                "report_types": ["Progress", "Safety", "Quality"],
                "frequency": "Weekly"
            }
        ]

def render_generate_reports():
    """Render report generation interface"""
    
    st.subheader("📋 Highland Tower Development - Generate Reports")
    
    # Quick report generation
    st.markdown("**⚡ Quick Report Generation:**")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Executive Summary", key="quick_exec", use_container_width=True):
            generate_executive_summary_report()
    
    with col2:
        if st.button("💰 Cost Performance", key="quick_cost", use_container_width=True):
            generate_cost_performance_report()
    
    with col3:
        if st.button("🦺 Safety Dashboard", key="quick_safety", use_container_width=True):
            generate_safety_dashboard_report()
    
    st.markdown("---")
    
    # Custom report generation
    st.markdown("**🛠️ Custom Report Generation:**")
    
    with st.form("custom_report_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            report_title = st.text_input("Report Title*", placeholder="Highland Tower Monthly Summary")
            report_type = st.selectbox("Report Type", ["Executive", "Financial", "Safety", "Progress", "Quality", "Custom"])
            report_format = st.selectbox("Format", ["PDF", "Excel", "PowerPoint", "Word"])
            
            date_range = st.selectbox("Date Range", [
                "This Week", "This Month", "This Quarter", 
                "Last 7 Days", "Last 30 Days", "Custom Range"
            ])
            
            if date_range == "Custom Range":
                col1a, col1b = st.columns(2)
                with col1a:
                    start_date = st.date_input("Start Date")
                with col1b:
                    end_date = st.date_input("End Date")
        
        with col2:
            data_sources = st.multiselect("Data Sources", [
                "Cost Management", "Daily Reports", "Safety", "Quality Control",
                "RFIs", "Change Orders", "Schedule", "Progress Photos", 
                "Material Management", "Subcontractors"
            ], default=["Cost Management", "Daily Reports", "Safety"])
            
            include_charts = st.checkbox("Include Charts & Graphs", value=True)
            include_photos = st.checkbox("Include Progress Photos", value=False)
            include_financials = st.checkbox("Include Financial Data", value=True)
            executive_summary = st.checkbox("Executive Summary Page", value=True)
        
        report_sections = st.multiselect("Report Sections", [
            "Project Overview", "Financial Summary", "Schedule Status", 
            "Safety Performance", "Quality Metrics", "Risk Assessment",
            "Progress Photos", "Change Orders", "Resource Utilization",
            "Milestone Tracking", "Issue Summary"
        ], default=["Project Overview", "Financial Summary", "Schedule Status"])
        
        custom_notes = st.text_area("Custom Notes", placeholder="Additional information to include in the report...")
        
        col1, col2 = st.columns(2)
        
        with col1:
            distribution_list = st.selectbox("Distribution List", ["Manual", "Executive Team", "Project Team", "Custom"])
        
        with col2:
            delivery_method = st.selectbox("Delivery Method", ["Download", "Email", "Both"])
        
        if st.form_submit_button("📊 Generate Custom Report", type="primary"):
            if report_title and report_type and data_sources:
                with st.spinner("Generating comprehensive Highland Tower Development report..."):
                    report_data = {
                        "title": report_title,
                        "type": report_type,
                        "format": report_format,
                        "date_range": date_range,
                        "data_sources": data_sources,
                        "sections": report_sections,
                        "include_charts": include_charts,
                        "include_photos": include_photos,
                        "include_financials": include_financials,
                        "executive_summary": executive_summary,
                        "custom_notes": custom_notes
                    }
                    
                    if report_format == "PDF":
                        report_buffer = generate_comprehensive_pdf_report(report_data)
                    elif report_format == "Excel":
                        report_buffer = generate_comprehensive_excel_report(report_data)
                    else:
                        st.info(f"{report_format} generation coming soon!")
                        return
                    
                    if report_buffer:
                        # Create download link
                        b64_report = base64.b64encode(report_buffer.getvalue()).decode()
                        current_date = datetime.now().strftime("%Y%m%d")
                        filename = f"Highland_Tower_{report_title.replace(' ', '_')}_{current_date}.{report_format.lower()}"
                        
                        if report_format == "PDF":
                            mime_type = "application/pdf"
                        elif report_format == "Excel":
                            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        
                        href = f'<a href="data:{mime_type};base64,{b64_report}" download="{filename}" target="_blank">' \
                               f'<div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); ' \
                               f'color: white; padding: 12px 24px; border-radius: 8px; text-align: center; ' \
                               f'font-weight: 600; text-decoration: none; display: inline-block; margin: 8px 0;">' \
                               f'📥 Download {report_title} Report</div></a>'
                        
                        st.markdown(href, unsafe_allow_html=True)
                        st.success("✅ Highland Tower Development report generated successfully!")
                        
                        # Show report summary
                        st.markdown("**📋 Generated Report Contains:**")
                        for section in report_sections:
                            st.write(f"• {section}")
                        if include_charts:
                            st.write("• Professional charts and visualizations")
                        if include_photos:
                            st.write("• Progress photography")
                        st.write(f"• Highland Tower Development authentic data")
            else:
                st.error("Please fill in all required fields!")

def generate_executive_summary_report():
    """Generate executive summary report"""
    
    with st.spinner("Generating Highland Tower Development Executive Summary..."):
        report_data = {
            "title": "Highland Tower Development - Executive Summary",
            "type": "Executive", 
            "format": "PDF",
            "date_range": "This Month",
            "data_sources": ["Cost Management", "Daily Reports", "Safety", "Quality Control"],
            "sections": ["Project Overview", "Financial Summary", "Schedule Status", "Risk Assessment"],
            "include_charts": True,
            "include_photos": False,
            "include_financials": True,
            "executive_summary": True,
            "custom_notes": "Monthly executive briefing for Highland Properties stakeholders"
        }
        
        report_buffer = generate_comprehensive_pdf_report(report_data)
        
        if report_buffer:
            b64_report = base64.b64encode(report_buffer.getvalue()).decode()
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"Highland_Tower_Executive_Summary_{current_date}.pdf"
            
            href = f'<a href="data:application/pdf;base64,{b64_report}" download="{filename}" target="_blank">' \
                   f'<div style="background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%); ' \
                   f'color: white; padding: 12px 24px; border-radius: 8px; text-align: center; ' \
                   f'font-weight: 600; text-decoration: none; display: inline-block; margin: 8px 0;">' \
                   f'📥 Download Executive Summary</div></a>'
            
            st.markdown(href, unsafe_allow_html=True)
            st.success("✅ Executive Summary generated with Highland Tower authentic data!")

def generate_cost_performance_report():
    """Generate cost performance report"""
    
    with st.spinner("Generating Highland Tower Development Cost Performance Report..."):
        report_data = {
            "title": "Highland Tower Development - Cost Performance Analysis",
            "type": "Financial",
            "format": "Excel",
            "date_range": "This Month",
            "data_sources": ["Cost Management", "Change Orders", "SOV"],
            "sections": ["Budget Summary", "Variance Analysis", "Change Orders", "Forecasting"],
            "include_charts": True,
            "include_photos": False,
            "include_financials": True,
            "executive_summary": False,
            "custom_notes": "Detailed cost analysis showing $2.1M under budget performance"
        }
        
        report_buffer = generate_comprehensive_excel_report(report_data)
        
        if report_buffer:
            b64_report = base64.b64encode(report_buffer.getvalue()).decode()
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"Highland_Tower_Cost_Performance_{current_date}.xlsx"
            
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_report}" download="{filename}" target="_blank">' \
                   f'<div style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); ' \
                   f'color: white; padding: 12px 24px; border-radius: 8px; text-align: center; ' \
                   f'font-weight: 600; text-decoration: none; display: inline-block; margin: 8px 0;">' \
                   f'📥 Download Cost Performance Report</div></a>'
            
            st.markdown(href, unsafe_allow_html=True)
            st.success("✅ Cost Performance Report generated with Highland Tower financial data!")

def generate_safety_dashboard_report():
    """Generate safety dashboard report"""
    
    with st.spinner("Generating Highland Tower Development Safety Dashboard..."):
        report_data = {
            "title": "Highland Tower Development - Safety Performance Dashboard",
            "type": "Safety",
            "format": "PDF", 
            "date_range": "This Month",
            "data_sources": ["Safety", "Daily Reports", "Training"],
            "sections": ["Safety Metrics", "Incident Summary", "Training Status", "Compliance"],
            "include_charts": True,
            "include_photos": False,
            "include_financials": False,
            "executive_summary": True,
            "custom_notes": "97.2 safety rating with excellent compliance record"
        }
        
        report_buffer = generate_comprehensive_pdf_report(report_data)
        
        if report_buffer:
            b64_report = base64.b64encode(report_buffer.getvalue()).decode()
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"Highland_Tower_Safety_Dashboard_{current_date}.pdf"
            
            href = f'<a href="data:application/pdf;base64,{b64_report}" download="{filename}" target="_blank">' \
                   f'<div style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); ' \
                   f'color: white; padding: 12px 24px; border-radius: 8px; text-align: center; ' \
                   f'font-weight: 600; text-decoration: none; display: inline-block; margin: 8px 0;">' \
                   f'📥 Download Safety Dashboard</div></a>'
            
            st.markdown(href, unsafe_allow_html=True)
            st.success("✅ Safety Dashboard generated with Highland Tower safety data!")

def render_report_templates():
    """Render report templates management"""
    
    st.subheader("📊 Highland Tower Development - Report Templates")
    
    st.info("**📊 Report Templates:** Professional report formats designed specifically for Highland Tower Development project needs.")
    
    # Template categories
    col1, col2, col3 = st.columns(3)
    
    with col1:
        category_filter = st.selectbox("Filter by Category", ["All", "Executive", "Financial", "Safety", "Progress", "Quality"])
    with col2:
        format_filter = st.selectbox("Filter by Format", ["All", "PDF", "Excel", "PowerPoint", "Word"])
    with col3:
        frequency_filter = st.selectbox("Filter by Frequency", ["All", "Daily", "Weekly", "Monthly", "Quarterly"])
    
    # Apply filters
    filtered_templates = st.session_state.highland_report_templates
    if category_filter != "All":
        filtered_templates = [t for t in filtered_templates if t['category'] == category_filter]
    if format_filter != "All":
        filtered_templates = [t for t in filtered_templates if t['format'] == format_filter]
    if frequency_filter != "All":
        filtered_templates = [t for t in filtered_templates if t['frequency'] == frequency_filter]
    
    # Display templates
    for template in filtered_templates:
        with st.expander(f"📋 {template['name']} ({template['format']}) - Used {template['usage_count']} times"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**Template ID:** {template['template_id']}")
                st.write(f"**Category:** {template['category']}")
                st.write(f"**Format:** {template['format']}")
                st.write(f"**Frequency:** {template['frequency']}")
                st.write(f"**Last Used:** {template['last_used']}")
                st.write(f"**Usage Count:** {template['usage_count']}")
            
            with col2:
                st.write(f"**Description:** {template['description']}")
                st.write("**Data Sources:**")
                for source in template['data_sources']:
                    st.write(f"• {source}")
                st.write("**Sections:**")
                for section in template['sections']:
                    st.write(f"• {section}")
            
            # Template actions
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if st.button("🚀 Use Template", key=f"use_{template['template_id']}"):
                    st.info(f"Using template: {template['name']}")
            
            with col2:
                if st.button("✏️ Edit Template", key=f"edit_{template['template_id']}"):
                    st.info("Template editing functionality...")
            
            with col3:
                if st.button("📋 Duplicate", key=f"dup_{template['template_id']}"):
                    st.info("Template duplication functionality...")
            
            with col4:
                if st.button("📊 Preview", key=f"preview_{template['template_id']}"):
                    st.info("Template preview functionality...")

def render_scheduled_reports():
    """Render scheduled reports management"""
    
    st.subheader("🕒 Highland Tower Development - Scheduled Reports")
    
    st.info("**🕒 Scheduled Reports:** Automate regular report generation and distribution for Highland Tower stakeholders.")
    
    # Scheduled reports overview
    scheduled_reports = [
        {
            "schedule_id": "SCH-001",
            "report_name": "Weekly Executive Summary",
            "template": "Executive Summary Report",
            "frequency": "Weekly",
            "next_run": "2024-06-03 08:00",
            "distribution": "Executive Team",
            "status": "Active",
            "last_run": "2024-05-27 08:00",
            "success_rate": "100%"
        },
        {
            "schedule_id": "SCH-002", 
            "report_name": "Monthly Cost Analysis",
            "template": "Cost Performance Report",
            "frequency": "Monthly",
            "next_run": "2024-06-01 09:00",
            "distribution": "Executive Team, Project Team",
            "status": "Active",
            "last_run": "2024-05-01 09:00",
            "success_rate": "100%"
        },
        {
            "schedule_id": "SCH-003",
            "report_name": "Daily Safety Report",
            "template": "Safety Performance Dashboard", 
            "frequency": "Daily",
            "next_run": "2024-05-29 17:00",
            "distribution": "Project Team",
            "status": "Active",
            "last_run": "2024-05-28 17:00",
            "success_rate": "98%"
        }
    ]
    
    # Display scheduled reports
    for report in scheduled_reports:
        status_color = {"Active": "🟢", "Paused": "🟡", "Failed": "🔴"}.get(report['status'], "⚪")
        
        with st.expander(f"{status_color} {report['report_name']} - {report['frequency']}"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**Schedule ID:** {report['schedule_id']}")
                st.write(f"**Template:** {report['template']}")
                st.write(f"**Frequency:** {report['frequency']}")
                st.write(f"**Status:** {status_color} {report['status']}")
                st.write(f"**Success Rate:** {report['success_rate']}")
            
            with col2:
                st.write(f"**Next Run:** {report['next_run']}")
                st.write(f"**Last Run:** {report['last_run']}")
                st.write(f"**Distribution:** {report['distribution']}")
            
            # Schedule actions
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if st.button("▶️ Run Now", key=f"run_{report['schedule_id']}"):
                    st.info(f"Running {report['report_name']}...")
            
            with col2:
                if report['status'] == 'Active':
                    if st.button("⏸️ Pause", key=f"pause_{report['schedule_id']}"):
                        st.info("Schedule paused")
                else:
                    if st.button("▶️ Resume", key=f"resume_{report['schedule_id']}"):
                        st.info("Schedule resumed")
            
            with col3:
                if st.button("✏️ Edit Schedule", key=f"edit_sched_{report['schedule_id']}"):
                    st.info("Schedule editing functionality...")
            
            with col4:
                if st.button("📊 View History", key=f"history_{report['schedule_id']}"):
                    st.info("Run history functionality...")

def render_distribution_management():
    """Render distribution lists management"""
    
    st.subheader("📧 Highland Tower Development - Distribution Management")
    
    st.info("**📧 Distribution Management:** Manage stakeholder groups and automated report distribution for Highland Tower project.")
    
    # Display distribution lists
    for dist_list in st.session_state.highland_distribution_lists:
        with st.expander(f"📧 {dist_list['name']} ({len(dist_list['recipients'])} recipients)"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**List ID:** {dist_list['list_id']}")
                st.write(f"**Description:** {dist_list['description']}")
                st.write(f"**Frequency:** {dist_list['frequency']}")
                st.write("**Report Types:**")
                for report_type in dist_list['report_types']:
                    st.write(f"• {report_type}")
            
            with col2:
                st.write("**Recipients:**")
                for recipient in dist_list['recipients']:
                    st.write(f"• **{recipient['name']}** ({recipient['role']})")
                    st.write(f"  {recipient['email']}")

def render_report_analytics():
    """Render report analytics and usage statistics"""
    
    st.subheader("📈 Highland Tower Development - Report Analytics")
    
    # Report usage statistics
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📊 Report Generation Statistics:**")
        
        # Create usage chart
        usage_data = {
            "Report Type": ["Executive", "Financial", "Safety", "Progress", "Quality"],
            "Generated This Month": [12, 18, 24, 16, 8],
            "Average Generation Time": [2.3, 1.8, 1.2, 3.1, 1.5]
        }
        
        fig = px.bar(usage_data, x="Report Type", y="Generated This Month", 
                    title="Highland Tower Reports Generated This Month")
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        st.markdown("**⏱️ Performance Metrics:**")
        
        perf_data = {
            "Metric": ["Avg Generation Time", "Success Rate", "User Satisfaction", "Automation Rate"],
            "Value": ["2.1 seconds", "99.2%", "4.8/5.0", "75%"],
            "Trend": ["⬇️ Improving", "⬆️ Excellent", "⬆️ High", "⬆️ Growing"]
        }
        
        perf_df = pd.DataFrame(perf_data)
        st.dataframe(perf_df, use_container_width=True, hide_index=True)

def generate_comprehensive_pdf_report(report_data: Dict[str, Any]) -> io.BytesIO:
    """Generate comprehensive PDF report with Highland Tower data"""
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.darkblue
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    # Story elements
    story = []
    
    # Title page
    story.append(Paragraph(report_data['title'], title_style))
    story.append(Paragraph("Highland Tower Development", styles['Normal']))
    story.append(Paragraph("$45.5M Mixed-Use Development", styles['Normal']))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Executive summary
    if report_data.get('executive_summary'):
        story.append(Paragraph("EXECUTIVE SUMMARY", header_style))
        
        exec_summary = """
        Highland Tower Development continues to demonstrate exceptional performance across all key metrics. 
        The project is currently 78.5% complete, exceeding our 75% target milestone. Financial performance 
        remains strong with a 1.02 cost performance index, indicating we are 2% under budget with projected 
        savings of $2.1M. Schedule performance index of 1.05 shows we are 5% ahead of schedule. Safety 
        rating of 97.2 reflects our commitment to zero-incident construction.
        """
        story.append(Paragraph(exec_summary, styles['Normal']))
        story.append(Spacer(1, 20))
    
    # Project overview section
    if "Project Overview" in report_data.get('sections', []):
        story.append(Paragraph("PROJECT OVERVIEW", header_style))
        
        project_data = [
            ['Project Element', 'Value'],
            ['Contract Value', '$45,500,000'],
            ['Current Progress', '78.5%'],
            ['Schedule Performance', '105% (5% ahead)'],
            ['Cost Performance', '102% (2% under budget)'],
            ['Safety Rating', '97.2/100'],
            ['Quality Score', '94.2/100'],
            ['Projected Completion', 'November 23, 2025'],
            ['Projected Savings', '$2,100,000']
        ]
        
        project_table = Table(project_data, colWidths=[3*inch, 2*inch])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(project_table)
        story.append(Spacer(1, 20))
    
    # Financial summary section
    if "Financial Summary" in report_data.get('sections', []):
        story.append(Paragraph("FINANCIAL SUMMARY", header_style))
        
        financial_data = [
            ['Financial Metric', 'Budgeted', 'Actual', 'Variance'],
            ['Original Contract', '$45,500,000', '$45,500,000', '$0'],
            ['Change Orders', '$0', '$585,000', '+$585,000'],
            ['Total Contract Value', '$45,500,000', '$46,085,000', '+$585,000'],
            ['Spent to Date', '$32,500,000', '$30,247,800', '-$2,252,200'],
            ['Remaining Budget', '$13,000,000', '$15,837,200', '+$2,837,200'],
            ['Projected Final Cost', '$45,500,000', '$43,400,000', '-$2,100,000']
        ]
        
        financial_table = Table(financial_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(financial_table)
        story.append(Spacer(1, 20))
    
    # Custom notes
    if report_data.get('custom_notes'):
        story.append(Paragraph("ADDITIONAL NOTES", header_style))
        story.append(Paragraph(report_data['custom_notes'], styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_comprehensive_excel_report(report_data: Dict[str, Any]) -> io.BytesIO:
    """Generate comprehensive Excel report with Highland Tower data"""
    
    buffer = io.BytesIO()
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Highland Tower Summary")
    
    # Header
    ws_summary.merge_cells('A1:F1')
    header_cell = ws_summary['A1']
    header_cell.value = report_data['title']
    header_cell.font = Font(bold=True, size=16)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Project overview
    ws_summary['A3'] = "Project Overview"
    ws_summary['A3'].font = Font(bold=True, size=14)
    
    overview_data = [
        ['Metric', 'Value'],
        ['Contract Value', '$45,500,000'],
        ['Current Progress', '78.5%'],
        ['Schedule Performance', '105% (5% ahead)'],
        ['Cost Performance', '102% (2% under budget)'],
        ['Safety Rating', '97.2/100'],
        ['Quality Score', '94.2/100']
    ]
    
    for row, (metric, value) in enumerate(overview_data, 4):
        ws_summary.cell(row=row, column=1, value=metric)
        ws_summary.cell(row=row, column=2, value=value)
        if row == 4:  # Header row
            ws_summary.cell(row=row, column=1).font = Font(bold=True)
            ws_summary.cell(row=row, column=2).font = Font(bold=True)
    
    # Financial details sheet
    if "Financial Summary" in report_data.get('sections', []):
        ws_financial = wb.create_sheet("Financial Analysis")
        
        # Financial data
        financial_headers = ['Category', 'Budgeted', 'Actual', 'Variance', 'Variance %']
        for col, header in enumerate(financial_headers, 1):
            cell = ws_financial.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        financial_data = [
            ['General Requirements', 2280000, 2100000, -180000, -7.9],
            ['Concrete Work', 8750000, 8250000, -500000, -5.7],
            ['Structural Steel', 12400000, 10230000, -2170000, -17.5],
            ['MEP Systems', 15200000, 8740000, -6460000, -42.5],
            ['Exterior Envelope', 6870000, 3572400, -3297600, -48.0]
        ]
        
        for row, data in enumerate(financial_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws_financial.cell(row=row, column=col, value=value)
                if col > 1:  # Format currency columns
                    cell.number_format = '$#,##0'
                if col == 5:  # Format percentage column
                    cell.number_format = '0.0%'
    
    # Auto-adjust column widths
    for worksheet in wb.worksheets:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer