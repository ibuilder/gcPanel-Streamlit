"""
AIA G702/G703 Billing Module for gcPanel

This module provides AIA G702/G703 billing functionality for the Cost Management section,
enabling creation, editing, and export of standard AIA payment applications.
"""

import streamlit as st
import pandas as pd
import os
import json
from datetime import datetime, timedelta
import random
import base64
from io import BytesIO
import re

from modules.crud_template import CrudModule
from components.digital_signature import render_digital_signature_section, get_signature_summary, validate_required_signatures
from assets.crud_styler import (
    apply_crud_styles, 
    render_form_actions, 
    render_crud_fieldset
)

class AIABillingModule(CrudModule):
    def __init__(self):
        """Initialize the AIA Billing module with configuration."""
        super().__init__(
            module_name="AIA Billing",
            data_file_path="data/cost_management/aia_billing.json",
            id_field="payment_app_id",
            list_columns=["payment_app_id", "period_to", "total_completed", "payment_due", "status"],
            default_sort_field="period_to",
            default_sort_direction="desc",
            status_field="status",
            filter_options=["Draft", "Submitted", "Approved", "Paid"]
        )
        
        # Ensure data directory exists
        os.makedirs(os.path.dirname(self.data_file_path), exist_ok=True)
        
        # Initialize demo data if it doesn't exist
        if not os.path.exists(self.data_file_path) or len(self._get_items()) == 0:
            self._initialize_demo_data()
    
    def _initialize_demo_data(self):
        """Initialize sample data if none exists."""
        statuses = ["Draft", "Submitted", "Approved", "Paid"]
        
        demo_items = []
        
        # Create several sample payment applications
        for i in range(1, 6):
            # Set period dates
            period_to = (datetime.now() - timedelta(days=30*(5-i))).replace(day=25)
            period_from = (period_to.replace(day=1) - timedelta(days=1)).replace(day=26)
            
            # Create the schedule of values
            schedule_of_values = []
            total_contract_sum = 45500000.00  # From project info
            
            # Division percentages (should sum to 1.0)
            division_percentages = {
                "Division 01 - General Requirements": 0.07,
                "Division 02 - Site Construction": 0.08,
                "Division 03 - Concrete": 0.15,
                "Division 04 - Masonry": 0.09,
                "Division 05 - Metals": 0.12,
                "Division 06 - Wood & Plastics": 0.05,
                "Division 07 - Thermal & Moisture": 0.08,
                "Division 08 - Doors & Windows": 0.06,
                "Division 09 - Finishes": 0.07,
                "Division 10 - Specialties": 0.03,
                "Division 11 - Equipment": 0.04,
                "Division 12 - Furnishings": 0.03,
                "Division 13 - Special Construction": 0.05,
                "Division 14 - Conveying Systems": 0.04,
                "Division 21-23 - Mechanical": 0.15,
                "Division 26-28 - Electrical": 0.12
            }
            
            # Add the divisions to the schedule of values
            total_scheduled_value = 0
            item_id = 1
            
            for division, percentage in division_percentages.items():
                scheduled_value = round(total_contract_sum * percentage, 2)
                total_scheduled_value += scheduled_value
                
                # Set completion percentages based on application number and division
                # Earlier applications have less completion
                max_completion_pct = min(1.0, (i * 0.25) * (1 + random.uniform(-0.2, 0.2)))
                
                # Adjust completion percentage by division (earlier divisions complete faster)
                import re
                division_match = re.search(r'Division (\d+)', division)
                division_num = int(division_match.group(1)) if division_match else i
                division_factor = max(0.1, min(1.0, 1.0 - (division_num / 30)))
                
                completion_pct = max_completion_pct * division_factor
                
                # Calculate values for this period and previous periods
                if i == 1:
                    # First application
                    previous_applications = 0
                    this_period = round(scheduled_value * completion_pct, 2)
                else:
                    # Calculate from previous apps with increasing completion
                    prev_app_pct = min(1.0, ((i-1) * 0.25) * (1 + random.uniform(-0.2, 0.2))) * division_factor
                    previous_applications = round(scheduled_value * prev_app_pct, 2)
                    this_period = round((scheduled_value * completion_pct) - previous_applications, 2)
                
                # Ensure this_period is not negative
                this_period = max(0, this_period)
                
                # Calculate completed to date
                completed_to_date = previous_applications + this_period
                
                # Calculate balance to finish and percent complete
                balance_to_finish = scheduled_value - completed_to_date
                percent_complete = (completed_to_date / scheduled_value) * 100 if scheduled_value > 0 else 0
                
                # Add stored materials (occasionally)
                stored_materials = 0
                if random.random() < 0.2:  # 20% chance of having stored materials
                    stored_materials = round(scheduled_value * random.uniform(0.05, 0.15), 2)
                
                # Calculate total completed and stored
                total_completed_stored = completed_to_date + stored_materials
                
                schedule_item = {
                    "item_id": item_id,
                    "description": division,
                    "scheduled_value": scheduled_value,
                    "previous_applications": previous_applications,
                    "this_period": this_period,
                    "completed_to_date": completed_to_date,
                    "percent_complete": percent_complete,
                    "balance_to_finish": balance_to_finish,
                    "stored_materials": stored_materials,
                    "total_completed_stored": total_completed_stored
                }
                
                schedule_of_values.append(schedule_item)
                item_id += 1
            
            # Adjust the total scheduled value to match contract sum
            if total_scheduled_value != total_contract_sum:
                # Adjust the last item to make the total match
                difference = total_contract_sum - total_scheduled_value
                last_item = schedule_of_values[-1]
                last_item["scheduled_value"] += difference
                last_item["balance_to_finish"] += difference
            
            # Calculate totals
            total_scheduled = sum(item["scheduled_value"] for item in schedule_of_values)
            total_previous = sum(item["previous_applications"] for item in schedule_of_values)
            total_this_period = sum(item["this_period"] for item in schedule_of_values)
            total_completed = sum(item["completed_to_date"] for item in schedule_of_values)
            total_stored_materials = sum(item["stored_materials"] for item in schedule_of_values)
            total_completed_stored = sum(item["total_completed_stored"] for item in schedule_of_values)
            
            # Calculate retainage
            retainage_percentage = 5.0
            retainage_amount = round(total_completed_stored * (retainage_percentage / 100), 2)
            
            # Calculate total earned less retainage
            total_earned_less_retainage = total_completed_stored - retainage_amount
            
            # Calculate previous payments
            previous_payments = total_previous - (total_previous * (retainage_percentage / 100))
            
            # Calculate payment due
            payment_due = total_earned_less_retainage - previous_payments
            
            # Create the payment application
            payment_app = {
                "payment_app_id": f"PA-{i}",
                "application_number": i,
                "period_from": period_from.strftime('%Y-%m-%d'),
                "period_to": period_to.strftime('%Y-%m-%d'),
                "project_name": "Highland Tower Development",
                "owner_name": "Highland Properties LLC",
                "contractor_name": "GC Construction, Inc.",
                "architect_name": "Modern Architects Group",
                "contract_date": "2025-01-15",
                "project_number": "HT-2025-001",
                "contract_number": "OC-001",
                "total_contract_sum": total_contract_sum,
                "net_change_orders": 0.00,  # Assuming no change orders for the sample
                "contract_sum_to_date": total_contract_sum,
                "total_completed": total_completed,
                "total_stored_materials": total_stored_materials,
                "total_completed_stored": total_completed_stored,
                "retainage_percentage": retainage_percentage,
                "retainage_amount": retainage_amount,
                "total_earned_less_retainage": total_earned_less_retainage,
                "previous_payments": previous_payments,
                "payment_due": payment_due,
                "status": statuses[min(i-1, len(statuses)-1)],
                "date_created": (datetime.now() - timedelta(days=random.randint(1, 10))).strftime('%Y-%m-%d'),
                "created_by": "John Smith",
                "date_submitted": None if i == 1 else (datetime.now() - timedelta(days=random.randint(1, 5))).strftime('%Y-%m-%d'),
                "submitted_by": None if i == 1 else "John Smith",
                "date_approved": None if i <= 2 else (datetime.now() - timedelta(days=random.randint(1, 3))).strftime('%Y-%m-%d'),
                "approved_by": None if i <= 2 else "Jane Doe",
                "schedule_of_values": schedule_of_values
            }
            
            demo_items.append(payment_app)
        
        # Save demo data
        self._save_items(demo_items)
    
    def _create_new_item_template(self):
        """Create a template for a new payment application with default values."""
        payment_app_id = self._generate_new_id()
        
        # Get last payment app to determine the next application number
        items = sorted(self._get_items(), key=lambda x: x.get('application_number', 0))
        next_app_number = 1
        if items:
            next_app_number = items[-1].get('application_number', 0) + 1
        
        # Set period dates
        today = datetime.now()
        period_to = today.replace(day=min(today.day, 25))
        period_from = (period_to.replace(day=1) - timedelta(days=1)).replace(day=26)
        
        # Create blank schedule of values
        division_names = [
            "Division 01 - General Requirements", 
            "Division 02 - Site Construction", 
            "Division 03 - Concrete", 
            "Division 04 - Masonry", 
            "Division 05 - Metals",
            "Division 06 - Wood & Plastics", 
            "Division 07 - Thermal & Moisture", 
            "Division 08 - Doors & Windows", 
            "Division 09 - Finishes", 
            "Division 10 - Specialties",
            "Division 11 - Equipment", 
            "Division 12 - Furnishings", 
            "Division 13 - Special Construction", 
            "Division 14 - Conveying Systems", 
            "Division 21-23 - Mechanical",
            "Division 26-28 - Electrical"
        ]
        
        schedule_of_values = []
        for i, division in enumerate(division_names, 1):
            schedule_item = {
                "item_id": i,
                "description": division,
                "scheduled_value": 0.00,
                "previous_applications": 0.00,
                "this_period": 0.00,
                "completed_to_date": 0.00,
                "percent_complete": 0.00,
                "balance_to_finish": 0.00,
                "stored_materials": 0.00,
                "total_completed_stored": 0.00
            }
            schedule_of_values.append(schedule_item)
        
        return {
            "payment_app_id": f"PA-{payment_app_id}",
            "application_number": next_app_number,
            "period_from": period_from.strftime('%Y-%m-%d'),
            "period_to": period_to.strftime('%Y-%m-%d'),
            "project_name": "Highland Tower Development",
            "owner_name": "Highland Properties LLC",
            "contractor_name": "GC Construction, Inc.",
            "architect_name": "Modern Architects Group",
            "contract_date": "2025-01-15",
            "project_number": "HT-2025-001",
            "contract_number": "OC-001",
            "total_contract_sum": 45500000.00,
            "net_change_orders": 0.00,
            "contract_sum_to_date": 45500000.00,
            "total_completed": 0.00,
            "total_stored_materials": 0.00,
            "total_completed_stored": 0.00,
            "retainage_percentage": 5.0,
            "retainage_amount": 0.00,
            "total_earned_less_retainage": 0.00,
            "previous_payments": 0.00,
            "payment_due": 0.00,
            "status": "Draft",
            "date_created": datetime.now().strftime('%Y-%m-%d'),
            "created_by": "Current User",
            "date_submitted": None,
            "submitted_by": None,
            "date_approved": None,
            "approved_by": None,
            "schedule_of_values": schedule_of_values
        }
    
    def _generate_new_id(self):
        """Generate a new unique ID for payment applications."""
        items = self._get_items()
        
        # If no items exist yet, start with ID 1
        if not items:
            return "1"
        
        # Find the highest numerical ID and increment
        max_id = 0
        for item in items:
            payment_app_id = item.get('payment_app_id', '')
            if payment_app_id.startswith('PA-'):
                try:
                    num = int(payment_app_id.split('-')[1])
                    max_id = max(max_id, num)
                except:
                    pass
        
        return str(max_id + 1)
    
    def _get_status_class(self, status):
        """Get the status class for a given status value."""
        status_classes = {
            'Draft': 'secondary',
            'Submitted': 'info',
            'Approved': 'primary',
            'Paid': 'success'
        }
        return status_classes.get(status, 'secondary')
    
    def _recalculate_totals(self, item):
        """Recalculate all totals in the payment application."""
        # Calculate schedule of values totals
        total_scheduled = sum(sov_item["scheduled_value"] for sov_item in item["schedule_of_values"])
        total_previous = sum(sov_item["previous_applications"] for sov_item in item["schedule_of_values"])
        total_this_period = sum(sov_item["this_period"] for sov_item in item["schedule_of_values"])
        total_completed = sum(sov_item["completed_to_date"] for sov_item in item["schedule_of_values"])
        total_stored_materials = sum(sov_item["stored_materials"] for sov_item in item["schedule_of_values"])
        total_completed_stored = sum(sov_item["total_completed_stored"] for sov_item in item["schedule_of_values"])
        
        # Calculate retainage
        retainage_percentage = item["retainage_percentage"]
        retainage_amount = round(total_completed_stored * (retainage_percentage / 100), 2)
        
        # Calculate total earned less retainage
        total_earned_less_retainage = total_completed_stored - retainage_amount
        
        # Calculate previous payments (assuming previous payment apps determined this)
        previous_payments = item["previous_payments"]
        
        # Calculate payment due
        payment_due = total_earned_less_retainage - previous_payments
        
        # Update the item
        item["total_completed"] = total_completed
        item["total_stored_materials"] = total_stored_materials
        item["total_completed_stored"] = total_completed_stored
        item["retainage_amount"] = retainage_amount
        item["total_earned_less_retainage"] = total_earned_less_retainage
        item["payment_due"] = payment_due
        
        return item

    def render_detail_view(self):
        """Render the detail view for creating, viewing, or editing a payment application."""
        # Apply CRUD styles
        apply_crud_styles()
        
        base_key = self._get_state_key_prefix()
        
        # Get view mode
        is_new = st.session_state.get(f'{base_key}_view') == 'new'
        is_edit_mode = st.session_state.get(f'{base_key}_edit_mode', False)
        
        # Get item data
        if is_new:
            item = self._create_new_item_template()
            detail_title = f"New Payment Application #{item['application_number']}"
        else:
            item_id = st.session_state.get(f'{base_key}_selected_id')
            item = self._get_item_by_id(item_id)
            if not item:
                st.error(f"Payment Application with ID {item_id} not found")
                return
            detail_title = f"Payment Application #{item['application_number']}"
        
        # Render the detail container
        from assets.crud_styler import render_crud_detail_container, end_crud_detail_container
        
        mode_prefix = "New" if is_new else "Edit" if is_edit_mode else "View"
        container_title = f"{mode_prefix}: {detail_title}"
        
        detail_actions = render_crud_detail_container(
            title=container_title,
            is_new=is_new,
            back_button=True
        )
        
        # Check if back button was clicked
        if detail_actions['back_clicked']:
            st.session_state[f'{base_key}_view'] = 'list'
            st.rerun()
        
        # Display top action buttons for view mode
        if not is_new and not is_edit_mode:
            col1, col2, col3, col4, col5 = st.columns([1, 1, 1, 1, 4])
            with col1:
                if st.button("✏️ Edit", type="primary", key=f"edit_{base_key}_action"):
                    st.session_state[f'{base_key}_edit_mode'] = True
                    st.rerun()
            with col2:
                if st.button("📄 PDF", type="secondary", key=f"pdf_{base_key}_action"):
                    st.session_state[f'{base_key}_show_pdf'] = True
            with col3:
                if st.button("📊 Excel", type="secondary", key=f"excel_{base_key}_action"):
                    excel_bytes = self._generate_excel(item)
                    st.download_button(
                        label="Download Excel",
                        data=excel_bytes,
                        file_name=f"Payment_App_{item['application_number']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_excel_{base_key}"
                    )
            with col4:
                if st.button("🗑️ Delete", type="secondary", key=f"delete_{base_key}_action"):
                    st.warning("Are you sure you want to delete this payment application?")
                    confirm_col1, confirm_col2 = st.columns(2)
                    with confirm_col1:
                        if st.button("Yes, Delete", type="secondary", key=f"confirm_delete_{base_key}"):
                            self._delete_item(item['payment_app_id'])
                            st.success("Payment application deleted successfully")
                            st.session_state[f'{base_key}_view'] = 'list'
                            st.rerun()
                    with confirm_col2:
                        if st.button("Cancel", key=f"cancel_delete_{base_key}"):
                            st.rerun()
        
        # Check if we should show PDF
        if st.session_state.get(f'{base_key}_show_pdf', False):
            self._show_pdf(item)
            if st.button("Close PDF", key=f"close_pdf_{base_key}"):
                st.session_state[f'{base_key}_show_pdf'] = False
                st.rerun()
            return
        
        # Create form for editing or read-only view for viewing
        if is_edit_mode or is_new:
            with st.form(f"{base_key}_form"):
                # Basic Information Section
                def render_basic_info():
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        app_number = st.number_input("Application Number", 
                                                  value=int(item['application_number']), 
                                                  min_value=1,
                                                  step=1,
                                                  disabled=not is_new)
                    with col2:
                        period_from = st.date_input("Period From", 
                                              value=datetime.strptime(item['period_from'], '%Y-%m-%d'))
                    with col3:
                        period_to = st.date_input("Period To", 
                                            value=datetime.strptime(item['period_to'], '%Y-%m-%d'))
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        project_name = st.text_input("Project Name", value=item['project_name'])
                    with col2:
                        project_number = st.text_input("Project Number", value=item['project_number'])
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        owner_name = st.text_input("Owner", value=item['owner_name'])
                    with col2:
                        architect_name = st.text_input("Architect", value=item['architect_name'])
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        contractor_name = st.text_input("Contractor", value=item['contractor_name'])
                    with col2:
                        contract_number = st.text_input("Contract Number", value=item['contract_number'])
                    with col3:
                        contract_date = st.date_input("Contract Date", 
                                                value=datetime.strptime(item['contract_date'], '%Y-%m-%d'))
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        total_contract_sum = st.number_input("Original Contract Sum", 
                                                        value=float(item['total_contract_sum']),
                                                        min_value=0.0,
                                                        step=10000.0,
                                                        format="%.2f")
                    with col2:
                        net_change_orders = st.number_input("Net Change By Change Orders", 
                                                       value=float(item['net_change_orders']),
                                                       step=1000.0,
                                                       format="%.2f")
                    with col3:
                        # Calculate contract sum to date
                        contract_sum_to_date = total_contract_sum + net_change_orders
                        st.markdown(f"**Contract Sum To Date:** ${contract_sum_to_date:,.2f}")
                    
                    status = st.selectbox("Status", options=[
                        'Draft', 'Submitted', 'Approved', 'Paid'
                    ], index=['Draft', 'Submitted', 'Approved', 'Paid'].index(
                        item['status'] if item['status'] in ['Draft', 'Submitted', 'Approved', 'Paid'] 
                        else 'Draft')
                    )
                    
                    # Update fields based on status
                    if status == 'Submitted' and not item.get('date_submitted'):
                        date_submitted = datetime.now().strftime('%Y-%m-%d')
                        submitted_by = "Current User"
                    else:
                        date_submitted = item.get('date_submitted')
                        submitted_by = item.get('submitted_by')
                    
                    if status == 'Approved' and not item.get('date_approved'):
                        date_approved = datetime.now().strftime('%Y-%m-%d')
                        approved_by = "Current User"
                    else:
                        date_approved = item.get('date_approved')
                        approved_by = item.get('approved_by')
                
                render_crud_fieldset("Basic Information", render_basic_info)
                
                # Schedule of Values Section
                def render_schedule_of_values():
                    st.markdown("### Schedule of Values (G703)")
                    
                    # Display a table with editable fields
                    sov_data = item['schedule_of_values']
                    
                    # Create a header row
                    cols = st.columns([1, 4, 2, 2, 2, 2, 1, 2, 2, 2])
                    headers = ["#", "Description", "Scheduled Value", "Previous Applications", 
                              "This Period", "Completed To Date", "% Complete", 
                              "Balance To Finish", "Stored Materials", "Total Completed & Stored"]
                    
                    for i, header in enumerate(headers):
                        with cols[i]:
                            st.markdown(f"**{header}**")
                    
                    # Display each row with editable fields
                    updated_sov = []
                    
                    for index, sov_item in enumerate(sov_data):
                        cols = st.columns([1, 4, 2, 2, 2, 2, 1, 2, 2, 2])
                        
                        # Fixed fields
                        with cols[0]:
                            st.markdown(f"{sov_item['item_id']}")
                        
                        # Editable fields
                        with cols[1]:
                            description = st.text_input(f"Description {index}", 
                                                    value=sov_item['description'], 
                                                    label_visibility="collapsed")
                        
                        with cols[2]:
                            scheduled_value = st.number_input(f"Scheduled {index}", 
                                                         value=float(sov_item['scheduled_value']), 
                                                         min_value=0.0, 
                                                         step=1000.0, 
                                                         format="%.2f", 
                                                         label_visibility="collapsed")
                        
                        with cols[3]:
                            previous_applications = st.number_input(f"Previous {index}", 
                                                             value=float(sov_item['previous_applications']), 
                                                             min_value=0.0, 
                                                             max_value=float(scheduled_value),
                                                             step=1000.0, 
                                                             format="%.2f", 
                                                             label_visibility="collapsed")
                        
                        with cols[4]:
                            # This period should not exceed scheduled value - previous applications
                            max_this_period = max(0, scheduled_value - previous_applications)
                            this_period = st.number_input(f"This Period {index}", 
                                                     value=min(float(sov_item['this_period']), max_this_period), 
                                                     min_value=0.0, 
                                                     max_value=max_this_period,
                                                     step=1000.0, 
                                                     format="%.2f", 
                                                     label_visibility="collapsed")
                        
                        # Calculate the remaining fields
                        completed_to_date = previous_applications + this_period
                        percent_complete = (completed_to_date / scheduled_value) * 100 if scheduled_value > 0 else 0
                        balance_to_finish = scheduled_value - completed_to_date
                        
                        with cols[5]:
                            st.markdown(f"${completed_to_date:,.2f}")
                        
                        with cols[6]:
                            st.markdown(f"{percent_complete:.1f}%")
                        
                        with cols[7]:
                            st.markdown(f"${balance_to_finish:,.2f}")
                        
                        with cols[8]:
                            stored_materials = st.number_input(f"Stored {index}", 
                                                         value=float(sov_item['stored_materials']), 
                                                         min_value=0.0, 
                                                         step=1000.0, 
                                                         format="%.2f", 
                                                         label_visibility="collapsed")
                        
                        total_completed_stored = completed_to_date + stored_materials
                        
                        with cols[9]:
                            st.markdown(f"${total_completed_stored:,.2f}")
                        
                        # Update the SOV item
                        updated_sov_item = {
                            "item_id": sov_item['item_id'],
                            "description": description,
                            "scheduled_value": scheduled_value,
                            "previous_applications": previous_applications,
                            "this_period": this_period,
                            "completed_to_date": completed_to_date,
                            "percent_complete": percent_complete,
                            "balance_to_finish": balance_to_finish,
                            "stored_materials": stored_materials,
                            "total_completed_stored": total_completed_stored
                        }
                        
                        updated_sov.append(updated_sov_item)
                    
                    # Calculate totals
                    total_scheduled = sum(item["scheduled_value"] for item in updated_sov)
                    total_previous = sum(item["previous_applications"] for item in updated_sov)
                    total_this_period = sum(item["this_period"] for item in updated_sov)
                    total_completed = sum(item["completed_to_date"] for item in updated_sov)
                    total_percent = (total_completed / total_scheduled) * 100 if total_scheduled > 0 else 0
                    total_balance = total_scheduled - total_completed
                    total_stored_materials = sum(item["stored_materials"] for item in updated_sov)
                    total_completed_stored = sum(item["total_completed_stored"] for item in updated_sov)
                    
                    # Add a totals row
                    st.markdown("---")
                    cols = st.columns([1, 4, 2, 2, 2, 2, 1, 2, 2, 2])
                    
                    with cols[0]:
                        st.markdown("**#**")
                    with cols[1]:
                        st.markdown("**TOTALS**")
                    with cols[2]:
                        st.markdown(f"**${total_scheduled:,.2f}**")
                    with cols[3]:
                        st.markdown(f"**${total_previous:,.2f}**")
                    with cols[4]:
                        st.markdown(f"**${total_this_period:,.2f}**")
                    with cols[5]:
                        st.markdown(f"**${total_completed:,.2f}**")
                    with cols[6]:
                        st.markdown(f"**{total_percent:.1f}%**")
                    with cols[7]:
                        st.markdown(f"**${total_balance:,.2f}**")
                    with cols[8]:
                        st.markdown(f"**${total_stored_materials:,.2f}**")
                    with cols[9]:
                        st.markdown(f"**${total_completed_stored:,.2f}**")
                    
                    # Return the updated SOV for saving
                    return updated_sov, total_completed, total_stored_materials, total_completed_stored, contract_sum_to_date
                
                updated_sov, total_completed, total_stored_materials, total_completed_stored, contract_sum_to_date = render_schedule_of_values()
                
                # Summary Section (G702)
                def render_summary():
                    st.markdown("### Application Summary (G702)")
                    
                    # Calculate retainage
                    col1, col2 = st.columns(2)
                    with col1:
                        retainage_percentage = st.number_input("Retainage Percentage", 
                                                         value=float(item['retainage_percentage']),
                                                         min_value=0.0,
                                                         max_value=10.0,
                                                         step=0.5,
                                                         format="%.1f")
                    
                    retainage_amount = round(total_completed_stored * (retainage_percentage / 100), 2)
                    total_earned_less_retainage = total_completed_stored - retainage_amount
                    
                    with col2:
                        previous_payments = st.number_input("Previous Certificates for Payment", 
                                                      value=float(item['previous_payments']),
                                                      min_value=0.0,
                                                      step=1000.0,
                                                      format="%.2f")
                    
                    payment_due = total_earned_less_retainage - previous_payments
                    
                    st.markdown("---")
                    st.markdown("### Application Summary")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"**1. Original Contract Sum:** ${total_contract_sum:,.2f}")
                        st.markdown(f"**2. Net Change by Change Orders:** ${net_change_orders:,.2f}")
                        st.markdown(f"**3. Contract Sum to Date:** ${contract_sum_to_date:,.2f}")
                        st.markdown(f"**4. Total Completed & Stored to Date:** ${total_completed_stored:,.2f}")
                        st.markdown(f"**5. Retainage:**")
                        st.markdown(f"   **a. {retainage_percentage}% of Completed Work:** ${retainage_amount:,.2f}")
                    
                    with col2:
                        st.markdown(f"**6. Total Earned Less Retainage:** ${total_earned_less_retainage:,.2f}")
                        st.markdown(f"**7. Less Previous Certificates For Payment:** ${previous_payments:,.2f}")
                        st.markdown(f"**8. Current Payment Due:** ${payment_due:,.2f}")
                        st.markdown(f"**9. Balance To Finish, Including Retainage:** ${(contract_sum_to_date - total_earned_less_retainage):,.2f}")
                    
                    # Return values for saving
                    return retainage_percentage, retainage_amount, total_earned_less_retainage, previous_payments, payment_due
                
                retainage_percentage, retainage_amount, total_earned_less_retainage, previous_payments, payment_due = render_summary()
                
                # Form Actions
                form_actions = render_form_actions(
                    save_label="Save Payment Application",
                    cancel_label="Cancel",
                    delete_label="Delete Payment Application",
                    show_delete=not is_new
                )
                
                if form_actions['save_clicked']:
                    # Validate signatures before saving
                    is_valid, message = validate_required_signatures(signatures, ["Project Manager", "Owner Representative"])
                    if not is_valid:
                        st.error(f"Cannot save payment application: {message}")
                    else:
                        # Update item with form values
                        updated_item = {
                            "payment_app_id": item['payment_app_id'],
                            "application_number": int(app_number),
                            "period_from": period_from.strftime('%Y-%m-%d'),
                            "period_to": period_to.strftime('%Y-%m-%d'),
                            "project_name": project_name,
                            "owner_name": owner_name,
                            "contractor_name": contractor_name,
                            "architect_name": architect_name,
                            "contract_date": contract_date.strftime('%Y-%m-%d'),
                            "project_number": project_number,
                            "contract_number": contract_number,
                            "total_contract_sum": float(total_contract_sum),
                            "net_change_orders": float(net_change_orders),
                            "contract_sum_to_date": float(contract_sum_to_date),
                            "total_completed": float(total_completed),
                            "total_stored_materials": float(total_stored_materials),
                            "total_completed_stored": float(total_completed_stored),
                            "retainage_percentage": float(retainage_percentage),
                            "retainage_amount": float(retainage_amount),
                            "total_earned_less_retainage": float(total_earned_less_retainage),
                            "previous_payments": float(previous_payments),
                            "payment_due": float(payment_due),
                            "status": status,
                            "date_created": item['date_created'],
                            "created_by": item['created_by'],
                            "date_submitted": date_submitted,
                            "submitted_by": submitted_by,
                            "date_approved": date_approved,
                            "approved_by": approved_by,
                            "schedule_of_values": updated_sov,
                            "signatures": get_signature_summary(signatures)
                        }
                        
                        # Save the updated item
                        self._save_item(updated_item)
                        
                        # Show success message and return to list view
                        st.success("Payment application saved successfully with digital signatures!")
                        st.session_state[f'{base_key}_view'] = 'list'
                        st.rerun()
                
                if form_actions['cancel_clicked']:
                    st.session_state[f'{base_key}_view'] = 'list'
                    st.rerun()
                
                if form_actions['delete_clicked'] and not is_new:
                    self._delete_item(item['payment_app_id'])
                    st.success("Payment application deleted successfully")
                    st.session_state[f'{base_key}_view'] = 'list'
                    st.rerun()
        else:
            # Read-only view
            # Basic Information Section
            st.markdown(f"<div style='background-color: white; padding: 15px; border-radius: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
            st.subheader("Basic Information")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"**Application #:** {item['application_number']}")
            with col2:
                st.markdown(f"**Period From:** {item['period_from']}")
            with col3:
                st.markdown(f"**Period To:** {item['period_to']}")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**Project:** {item['project_name']}")
            with col2:
                st.markdown(f"**Project Number:** {item['project_number']}")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**Owner:** {item['owner_name']}")
            with col2:
                st.markdown(f"**Architect:** {item['architect_name']}")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"**Contractor:** {item['contractor_name']}")
            with col2:
                st.markdown(f"**Contract #:** {item['contract_number']}")
            with col3:
                st.markdown(f"**Contract Date:** {item['contract_date']}")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"**Original Contract Sum:** ${float(item['total_contract_sum']):,.2f}")
            with col2:
                st.markdown(f"**Net Change Orders:** ${float(item['net_change_orders']):,.2f}")
            with col3:
                st.markdown(f"**Contract Sum to Date:** ${float(item['contract_sum_to_date']):,.2f}")
            
            col1, col2 = st.columns(2)
            with col1:
                status_html = f"""
                <div style="display: flex; align-items: center; background: transparent;">
                    <span class='crud-status crud-status-{self._get_status_class(item['status'])}' style="outline: none; box-shadow: none; border: none;">{item['status']}</span>
                </div>
                """
                st.markdown("**Status:**")
                st.markdown(status_html, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"**Created Date:** {item['date_created']}")
            
            if item.get('date_submitted'):
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"**Submitted Date:** {item.get('date_submitted')}")
                with col2:
                    st.markdown(f"**Submitted By:** {item.get('submitted_by')}")
            
            if item.get('date_approved'):
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"**Approved Date:** {item.get('date_approved')}")
                with col2:
                    st.markdown(f"**Approved By:** {item.get('approved_by')}")
            
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Summary Section
            st.markdown(f"<div style='background-color: white; padding: 15px; border-radius: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
            st.subheader("Application Summary (G702)")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**1. Original Contract Sum:** ${float(item['total_contract_sum']):,.2f}")
                st.markdown(f"**2. Net Change by Change Orders:** ${float(item['net_change_orders']):,.2f}")
                st.markdown(f"**3. Contract Sum to Date:** ${float(item['contract_sum_to_date']):,.2f}")
                st.markdown(f"**4. Total Completed & Stored to Date:** ${float(item['total_completed_stored']):,.2f}")
                st.markdown(f"**5. Retainage:** {float(item['retainage_percentage'])}% = ${float(item['retainage_amount']):,.2f}")
            
            with col2:
                st.markdown(f"**6. Total Earned Less Retainage:** ${float(item['total_earned_less_retainage']):,.2f}")
                st.markdown(f"**7. Less Previous Certificates For Payment:** ${float(item['previous_payments']):,.2f}")
                st.markdown(f"**8. Current Payment Due:** ${float(item['payment_due']):,.2f}")
                balance_to_finish = float(item['contract_sum_to_date']) - float(item['total_earned_less_retainage'])
                st.markdown(f"**9. Balance To Finish, Including Retainage:** ${balance_to_finish:,.2f}")
            
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Schedule of Values
            st.markdown(f"<div style='background-color: white; padding: 15px; border-radius: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
            st.subheader("Schedule of Values (G703)")
            
            # Create a DataFrame for the schedule of values
            sov_df = pd.DataFrame(item['schedule_of_values'])
            
            # Format columns appropriately
            display_df = sov_df.copy()
            
            # Format currency columns
            currency_cols = ['scheduled_value', 'previous_applications', 'this_period', 'completed_to_date', 
                           'balance_to_finish', 'stored_materials', 'total_completed_stored']
            for col in currency_cols:
                display_df[col] = display_df[col].apply(lambda x: f"${float(x):,.2f}")
            
            # Format percentage column
            display_df['percent_complete'] = display_df['percent_complete'].apply(lambda x: f"{float(x):.1f}%")
            
            # Rename columns for display
            display_df = display_df.rename(columns={
                'item_id': '#',
                'description': 'Description',
                'scheduled_value': 'Scheduled Value',
                'previous_applications': 'Previous Applications',
                'this_period': 'This Period',
                'completed_to_date': 'Completed To Date',
                'percent_complete': '% Complete',
                'balance_to_finish': 'Balance To Finish',
                'stored_materials': 'Stored Materials',
                'total_completed_stored': 'Total Completed & Stored'
            })
            
            # Set the index to the item_id
            display_df = display_df.set_index('#')
            
            # Display the table
            st.dataframe(display_df, use_container_width=True)
            
            # Add totals row
            st.markdown("**TOTALS:**")
            
            total_scheduled = sum(sov_item["scheduled_value"] for sov_item in item['schedule_of_values'])
            total_previous = sum(sov_item["previous_applications"] for sov_item in item['schedule_of_values'])
            total_this_period = sum(sov_item["this_period"] for sov_item in item['schedule_of_values'])
            total_completed = sum(sov_item["completed_to_date"] for sov_item in item['schedule_of_values'])
            total_percent = (total_completed / total_scheduled) * 100 if total_scheduled > 0 else 0
            total_balance = total_scheduled - total_completed
            total_stored_materials = sum(sov_item["stored_materials"] for sov_item in item['schedule_of_values'])
            total_completed_stored = sum(sov_item["total_completed_stored"] for sov_item in item['schedule_of_values'])
            
            totals_row = {
                'Scheduled Value': f"${total_scheduled:,.2f}",
                'Previous Applications': f"${total_previous:,.2f}",
                'This Period': f"${total_this_period:,.2f}",
                'Completed To Date': f"${total_completed:,.2f}",
                '% Complete': f"{total_percent:.1f}%",
                'Balance To Finish': f"${total_balance:,.2f}",
                'Stored Materials': f"${total_stored_materials:,.2f}",
                'Total Completed & Stored': f"${total_completed_stored:,.2f}"
            }
            
            totals_df = pd.DataFrame([totals_row])
            st.dataframe(totals_df, use_container_width=True)
            
            st.markdown("</div>", unsafe_allow_html=True)
        
        end_crud_detail_container()
    
    def _generate_excel(self, item):
        """Generate an Excel file for the payment application."""
        try:
            import io
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
            from openpyxl.utils import get_column_letter
            
            # Create a BytesIO object to store the Excel file
            output = io.BytesIO()
            
            # Create a pandas Excel writer with openpyxl
            writer = pd.ExcelWriter(output, engine='openpyxl')
            
            # Extract data for G702 summary
            g702_data = {
                'Field': [
                    'Application No.', 
                    'Application Date', 
                    'Period From', 
                    'Period To',
                    'Project Name',
                    'Project No.',
                    'Contract No.',
                    'Contract Date',
                    'Owner',
                    'Contractor',
                    'Architect',
                    '1. Original Contract Sum',
                    '2. Net Change by Change Orders',
                    '3. Contract Sum to Date (Line 1 ± 2)',
                    '4. Total Completed & Stored to Date',
                    '5. Retainage',
                    '   a. % of Completed Work',
                    '6. Total Earned Less Retainage',
                    '7. Less Previous Certificates for Payment',
                    '8. Current Payment Due',
                    '9. Balance to Finish, Including Retainage'
                ],
                'Value': [
                    item['application_number'],
                    item['period_to'],
                    item['period_from'],
                    item['period_to'],
                    item['project_name'],
                    item['project_number'],
                    item['contract_number'],
                    item['contract_date'],
                    item['owner_name'],
                    item['contractor_name'],
                    item['architect_name'],
                    f"${float(item['total_contract_sum']):,.2f}",
                    f"${float(item['net_change_orders']):,.2f}",
                    f"${float(item['contract_sum_to_date']):,.2f}",
                    f"${float(item['total_completed_stored']):,.2f}",
                    f"${float(item['retainage_amount']):,.2f}",
                    f"{float(item['retainage_percentage'])}%",
                    f"${float(item['total_earned_less_retainage']):,.2f}",
                    f"${float(item['previous_payments']):,.2f}",
                    f"${float(item['payment_due']):,.2f}",
                    f"${float(item['contract_sum_to_date']) - float(item['total_earned_less_retainage']):,.2f}"
                ]
            }
            
            # Create G702 dataframe
            g702_df = pd.DataFrame(g702_data)
            g702_df.to_excel(writer, sheet_name='G702 Summary', index=False)
            
            # Format G702 worksheet
            g702_sheet = writer.sheets['G702 Summary']
            g702_sheet.column_dimensions['A'].width = 50
            g702_sheet.column_dimensions['B'].width = 25
            
            # Create G703 dataframe from schedule of values
            g703_data = []
            for sov_item in item['schedule_of_values']:
                g703_data.append({
                    'Item No.': sov_item['item_id'],
                    'Description of Work': sov_item['description'],
                    'Scheduled Value': sov_item['scheduled_value'],
                    'Previous Applications': sov_item['previous_applications'],
                    'This Period': sov_item['this_period'],
                    'Completed To Date': sov_item['completed_to_date'],
                    'Percent Complete': sov_item['percent_complete'],
                    'Balance To Finish': sov_item['balance_to_finish'],
                    'Stored Materials': sov_item['stored_materials'],
                    'Total Completed & Stored': sov_item['total_completed_stored']
                })
            
            g703_df = pd.DataFrame(g703_data)
            
            # Add totals row
            totals_row = {
                'Item No.': '',
                'Description of Work': 'TOTALS',
                'Scheduled Value': sum(sov_item['scheduled_value'] for sov_item in item['schedule_of_values']),
                'Previous Applications': sum(sov_item['previous_applications'] for sov_item in item['schedule_of_values']),
                'This Period': sum(sov_item['this_period'] for sov_item in item['schedule_of_values']),
                'Completed To Date': sum(sov_item['completed_to_date'] for sov_item in item['schedule_of_values']),
                'Percent Complete': '',  # Leave percent blank in totals
                'Balance To Finish': sum(sov_item['balance_to_finish'] for sov_item in item['schedule_of_values']),
                'Stored Materials': sum(sov_item['stored_materials'] for sov_item in item['schedule_of_values']),
                'Total Completed & Stored': sum(sov_item['total_completed_stored'] for sov_item in item['schedule_of_values'])
            }
            
            # Append totals row
            g703_df = g703_df._append(totals_row, ignore_index=True)
            
            # Write G703 to Excel
            g703_df.to_excel(writer, sheet_name='G703 Schedule of Values', index=False)
            
            # Format G703 worksheet
            g703_sheet = writer.sheets['G703 Schedule of Values']
            g703_sheet.column_dimensions['A'].width = 10
            g703_sheet.column_dimensions['B'].width = 40
            
            # Format numeric columns (C-J)
            for col_idx in range(3, 11):
                col_letter = get_column_letter(col_idx)
                g703_sheet.column_dimensions[col_letter].width = 15
            
            # Save the Excel writer
                pass  # Context manager handles save
            
            # Return the BytesIO object
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error generating Excel file: {str(e)}")
            return None
    
    def _show_pdf(self, item):
        """Generate and display a proper PDF version of the AIA G702/G703 payment application."""
        st.markdown("### 📄 AIA G702/G703 Payment Application PDF")
        
        try:
            # Generate the PDF using ReportLab
            pdf_bytes = self._generate_aia_pdf(item)
            
            if pdf_bytes:
                # Create download button for the PDF
                st.download_button(
                    label="📥 Download AIA G702/G703 PDF",
                    data=pdf_bytes,
                    file_name=f"AIA_Payment_App_{item['application_number']}.pdf",
                    mime="application/pdf",
                    type="primary"
                )
                
                # Display PDF preview in browser
                import base64
                base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
                pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="800px" type="application/pdf"></iframe>'
                st.markdown(pdf_display, unsafe_allow_html=True)
            else:
                st.error("❌ Error generating PDF. Please try again.")
                
        except Exception as e:
            st.error(f"❌ Error generating PDF: {str(e)}")

    def _generate_aia_pdf(self, item):
        """Generate a professional AIA G702/G703 PDF using ReportLab."""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from io import BytesIO
        
        try:
            # Create PDF in memory
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=1*inch,
                bottomMargin=1*inch
            )
            
            # Get styles
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=12,
                alignment=TA_LEFT
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=6
            )
            
            # Content for the PDF
            content = []
            
            # === PAGE 1: AIA G702 APPLICATION FOR PAYMENT ===
            
            # Title
            content.append(Paragraph("APPLICATION AND CERTIFICATE FOR PAYMENT", title_style))
            content.append(Paragraph("AIA DOCUMENT G702", normal_style))
            content.append(Spacer(1, 20))
            
            # Project Information Table
            project_data = [
                ['TO OWNER:', item['owner_name'], 'APPLICATION NO.:', str(item['application_number'])],
                ['PROJECT:', item['project_name'], 'PERIOD TO:', item['period_to']],
                ['PROJECT NO.:', item['project_number'], 'CONTRACT DATE:', item['contract_date']],
                ['', '', '', ''],
                ['FROM CONTRACTOR:', item['contractor_name'], 'VIA ARCHITECT:', item['architect_name']],
                ['CONTRACT FOR:', 'Construction', 'CONTRACT NO.:', item['contract_number']]
            ]
            
            project_table = Table(project_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 1.5*inch])
            project_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ]))
            
            content.append(project_table)
            content.append(Spacer(1, 30))
            
            # Payment Summary Table
            summary_data = [
                ['1. ORIGINAL CONTRACT SUM', f"${float(item['total_contract_sum']):,.2f}"],
                ['2. NET CHANGE BY CHANGE ORDERS', f"${float(item['net_change_orders']):,.2f}"],
                ['3. CONTRACT SUM TO DATE (Line 1 ± 2)', f"${float(item['contract_sum_to_date']):,.2f}"],
                ['4. TOTAL COMPLETED & STORED TO DATE', f"${float(item['total_completed_stored']):,.2f}"],
                [f"5. RETAINAGE ({item['retainage_percentage']}% of Completed Work)", f"${float(item['retainage_amount']):,.2f}"],
                ['6. TOTAL EARNED LESS RETAINAGE', f"${float(item['total_earned_less_retainage']):,.2f}"],
                ['7. LESS PREVIOUS CERTIFICATES FOR PAYMENT', f"${float(item['previous_payments']):,.2f}"],
                ['8. CURRENT PAYMENT DUE', f"${float(item['payment_due']):,.2f}"],
                ['9. BALANCE TO FINISH, INCLUDING RETAINAGE', f"${float(item['contract_sum_to_date']) - float(item['total_earned_less_retainage']):,.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[5*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 7), (1, 7), 'Helvetica-Bold'),  # Highlight current payment due
                ('BACKGROUND', (0, 7), (1, 7), colors.yellow),
            ]))
            
            content.append(summary_table)
            content.append(Spacer(1, 30))
            
            # Signature Section
            signature_data = [
                ['CONTRACTOR CERTIFICATION:', ''],
                ['By:', ''],
                ['Date:', ''],
                ['', ''],
                ['ARCHITECT CERTIFICATION:', ''],
                ['By:', ''],
                ['Date:', ''],
                ['', ''],
                ['OWNER APPROVAL:', ''],
                ['By:', ''],
                ['Date:', '']
            ]
            
            signature_table = Table(signature_data, colWidths=[3*inch, 3.5*inch])
            signature_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
                ('BACKGROUND', (0, 4), (0, 4), colors.lightgrey),
                ('BACKGROUND', (0, 8), (0, 8), colors.lightgrey),
            ]))
            
            content.append(signature_table)
            
            # Page break for G703
            content.append(PageBreak())
            
            # === PAGE 2: AIA G703 CONTINUATION SHEET ===
            
            content.append(Paragraph("CONTINUATION SHEET", title_style))
            content.append(Paragraph("AIA DOCUMENT G703", normal_style))
            content.append(Spacer(1, 20))
            
            # Project header for G703
            g703_header = [
                ['PROJECT:', item['project_name'], 'APPLICATION NO.:', str(item['application_number'])],
                ['CONTRACTOR:', item['contractor_name'], 'PERIOD TO:', item['period_to']]
            ]
            
            g703_header_table = Table(g703_header, colWidths=[1.5*inch, 3*inch, 1.5*inch, 1*inch])
            g703_header_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ]))
            
            content.append(g703_header_table)
            content.append(Spacer(1, 20))
            
            # Schedule of Values Table
            sov_headers = [
                'Item', 'Description of Work', 'Scheduled\nValue', 'Previous\nApps', 'This\nPeriod', 
                'Completed\nTo Date', '%\nComplete', 'Balance\nTo Finish', 'Stored\nMaterials', 'Total Completed\n& Stored'
            ]
            
            sov_data = [sov_headers]
            
            # Add schedule of values items
            for sov_item in item['schedule_of_values']:
                row = [
                    str(sov_item['item_id']),
                    sov_item['description'][:30] + '...' if len(sov_item['description']) > 30 else sov_item['description'],
                    f"${sov_item['scheduled_value']:,.0f}",
                    f"${sov_item['previous_applications']:,.0f}",
                    f"${sov_item['this_period']:,.0f}",
                    f"${sov_item['completed_to_date']:,.0f}",
                    f"{sov_item['percent_complete']:.1f}%",
                    f"${sov_item['balance_to_finish']:,.0f}",
                    f"${sov_item['stored_materials']:,.0f}",
                    f"${sov_item['total_completed_stored']:,.0f}"
                ]
                sov_data.append(row)
            
            # Add totals row
            total_scheduled = sum(sov['scheduled_value'] for sov in item['schedule_of_values'])
            total_previous = sum(sov['previous_applications'] for sov in item['schedule_of_values'])
            total_this_period = sum(sov['this_period'] for sov in item['schedule_of_values'])
            total_completed = sum(sov['completed_to_date'] for sov in item['schedule_of_values'])
            total_balance = sum(sov['balance_to_finish'] for sov in item['schedule_of_values'])
            total_stored = sum(sov['stored_materials'] for sov in item['schedule_of_values'])
            total_completed_stored = sum(sov['total_completed_stored'] for sov in item['schedule_of_values'])
            
            totals_row = [
                '', 'TOTALS', f"${total_scheduled:,.0f}", f"${total_previous:,.0f}", 
                f"${total_this_period:,.0f}", f"${total_completed:,.0f}", '', 
                f"${total_balance:,.0f}", f"${total_stored:,.0f}", f"${total_completed_stored:,.0f}"
            ]
            sov_data.append(totals_row)
            
            # Create the table with appropriate column widths
            col_widths = [0.4*inch, 2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch]
            sov_table = Table(sov_data, colWidths=col_widths, repeatRows=1)
            
            sov_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Header row
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Totals row
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header bold
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Totals bold
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),  # Right align numbers
                ('ALIGN', (0, 0), (1, -1), 'LEFT'),  # Left align item and description
            ]))
            
            content.append(sov_table)
            
            # Build PDF
            doc.build(content)
            
            # Return PDF bytes
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            st.error(f"Error generating AIA PDF: {str(e)}")
            return None

def render():
    """Render the AIA Billing module."""
    # Create and render the AIA Billing module
    aia_billing = AIABillingModule()
    aia_billing.render()